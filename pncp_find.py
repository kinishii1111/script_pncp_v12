#!/usr/bin/env python3
"""CLI para agentes: PNCP últimas N horas, filtrado pelo dataset de nicho.

Uso:
  python3 pncp_find.py --dataset ~/Downloads/dataset_2025.xlsx --horas 24 --json
  python3 pncp_find.py --dataset dataset_2025.xlsx --horas 24 --xlsx /tmp/out.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import unicodedata
from datetime import datetime, timedelta, timezone
from pathlib import Path
from fetch_retry import fetch_com_retry

BASE = "https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao"
# pregão e, conc, presencial, dispensa, inexig — volume vs sinal
MODALIDADES = (6, 8, 7, 4, 5, 9)
TAMANHO = 50
PAUSA = 0.7
TZ = timezone(timedelta(hours=-3))

# ruído que o léxico pega (decibelímetro, odonto, etc.)
BLOQUEIO = (
    r"decibel",
    r"bafomet",
    r"alcoolem",
    r"odont",
    r"esfigmo",
    r"reometro",
    r"ginecolog",
    r"instrumentais de",
    r"opme",
    r"audiovisual",
)

# frases do nicho (dataset 2025: hidrometria / instrumentação água-gás)
FORTE = (
    r"macromedidor",
    r"macro[\s-]?medidor",
    r"hidrometr",
    r"hidrometro",
    r"medidor(?:es)? de vazao",
    r"medidor(?:es)? de fluxo",
    r"medidor(?:es)? eletromagnet",
    r"medidor(?:es)? ultrasson",
    r"transmissor(?:es)? de pressao",
    r"transmissor(?:es)? de nivel",
    r"sensor(?:es)? de nivel",
    r"sensor(?:es)? de pressao",
    r"pitometr",
    r"rotametro",
    r"clamp[\s-]?on",
    r"calha parshall",
    r"woltmann",
    r"unijato",
    r"multijato",
    r"instrumentacao",
    r"telemetria",
    r"medidor(?:es)? de pressao",
)


def fold(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "")
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.lower())


def _celula_ok(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if len(s) < 8:
        return ""
    if s.startswith("=") or "DUMMYFUNCTION" in s or "REGEXREPLACE" in s:
        return ""
    return s


def carregar_dataset(path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    textos = []
    rows = ws.iter_rows(values_only=True)
    next(rows, None)
    for row in rows:
        row = list(row) + [None, None]
        t = _celula_ok(row[1]) or _celula_ok(row[0])
        if t:
            textos.append(t)
    wb.close()
    return textos


def resolver_datasets(paths: list[str] | None) -> list[Path]:
    if paths:
        out = []
        for p in paths:
            q = Path(p).expanduser()
            if q.is_dir():
                out.extend(sorted(q.glob("dataset_*.xlsx")))
            else:
                out.append(q)
        return out
    aqui = Path(__file__).resolve().parent / "data" / "nicho"
    found = sorted(aqui.glob("dataset_*.xlsx"))
    if found:
        return found
    dl = Path.home() / "Downloads"
    return sorted(dl.glob("dataset_20*.xlsx"))


_SINAL = re.compile(
    r"vazao|medidor|pressao|nivel|hidrom|macromed|calibr|telemetr|"
    r"esgoto|saneamento|adutora|saae|\beta\b|\bete\b"
)


def extra_frases(textos: list[str]) -> list[str]:
    """Bigramas do gabarito que ainda falam do nicho (não 'registro de preços')."""
    bag: dict[str, int] = {}
    for t in textos:
        toks = re.findall(r"[a-z0-9]+", fold(t))
        toks = [x for x in toks if len(x) > 3]
        for a, b in zip(toks, toks[1:]):
            bg = f"{a} {b}"
            if _SINAL.search(bg):
                bag[bg] = bag.get(bg, 0) + 1
    return [k for k, n in bag.items() if n >= 3][:40]


def pontuar(objeto: str, extras: list[str]) -> tuple[int, list[str]]:
    t = fold(objeto)
    if any(re.search(p, t) for p in BLOQUEIO):
        return 0, ["bloqueio"]
    hits = []
    sc = 0
    for p in FORTE:
        if re.search(p, t):
            hits.append(p)
            sc += 3
    if sc == 0:
        return 0, []
    for bg in extras:
        if bg in t:
            hits.append(bg)
            sc += 1
    return sc, hits[:8]


def parse_dt(s: str | None) -> datetime | None:
    if not s:
        return None
    s = str(s)[:19]
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=TZ)
        except ValueError:
            continue
    return None


def yyyymmdd(d: datetime) -> str:
    return d.strftime("%Y%m%d")


def coletar(
    corte: datetime,
    pausa: float,
    mods: tuple[int, ...],
    uf: str | None = None,
) -> list[dict]:
    hoje = datetime.now(TZ)
    dias = sorted({yyyymmdd(corte), yyyymmdd(hoje)})
    visto: dict[str, dict] = {}
    for dia in dias:
        for mod in mods:
            pagina = 1
            total = 1
            while pagina <= total:
                params = {
                    "dataInicial": dia,
                    "dataFinal": dia,
                    "codigoModalidadeContratacao": mod,
                    "pagina": pagina,
                    "tamanhoPagina": TAMANHO,
                }
                if uf:
                    params["uf"] = uf
                ret = fetch_com_retry(BASE, params)
                if ret is None:
                    print(f"[warn] falha {dia} mod={mod} p={pagina}", file=sys.stderr)
                    break
                dados, total_paginas = ret
                total = int(total_paginas or 1)
                for it in dados or []:
                    key = it.get("numeroControlePNCP")
                    if key:
                        visto[key] = it
                print(
                    f"[coleta] {dia} mod={mod} p={pagina}/{total} +{len(dados or [])}",
                    file=sys.stderr,
                )
                if pagina >= total:
                    break
                pagina += 1
                time.sleep(pausa)
            time.sleep(pausa)
    return list(visto.values())


def enriquece(it: dict, sc: int, hits: list[str], pub: datetime) -> dict:
    u = it.get("unidadeOrgao") or {}
    o = it.get("orgaoEntidade") or {}
    cnpj = o.get("cnpj")
    ano = it.get("anoCompra")
    seq = it.get("sequencialCompra")
    val = it.get("valorTotalEstimado")
    return {
        "score": sc,
        "hits": hits,
        "numero_pncp": it.get("numeroControlePNCP"),
        "data_publicacao": pub.strftime("%Y-%m-%d %H:%M"),
        "data_encerramento": (it.get("dataEncerramentoProposta") or "")[:19],
        "modalidade": it.get("modalidadeNome"),
        "uf": u.get("ufSigla"),
        "municipio": u.get("municipioNome"),
        "orgao": o.get("razaoSocial"),
        "cnpj": cnpj,
        "objeto": re.sub(r"\s+", " ", it.get("objetoCompra") or ""),
        "valor_estimado": val,
        "link_pncp": f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{seq}",
        "link_origem": it.get("linkSistemaOrigem") or "",
    }


def gravar_xlsx(rows: list[dict], path: Path, meta: dict | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "revisao"
    cols = [
        "ok",
        "comentario",
        "score",
        "numero_pncp",
        "data_publicacao",
        "data_encerramento",
        "modalidade",
        "uf",
        "municipio",
        "orgao",
        "cnpj",
        "valor_estimado",
        "objeto",
        "link_pncp",
        "hits",
    ]
    hdr_fill = PatternFill("solid", "1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    ws.append(cols)
    for c, _ in enumerate(cols, 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for r in rows:
        line = []
        for c in cols:
            if c == "ok":
                line.append("")
            elif c == "comentario":
                line.append("")
            elif c == "hits":
                line.append(", ".join(r.get("hits") or []))
            else:
                line.append(r.get(c))
        ws.append(line)
    for i in range(2, len(rows) + 2):
        ws.row_dimensions[i].height = 48
        for c in range(1, len(cols) + 1):
            ws.cell(i, c).alignment = wrap
            if cols[c - 1] == "link_pncp" and ws.cell(i, c).value:
                ws.cell(i, c).hyperlink = str(ws.cell(i, c).value)
                ws.cell(i, c).font = Font(color="0563C1", underline="single")
    widths = {"ok": 8, "comentario": 28, "objeto": 70, "orgao": 36, "link_pncp": 40, "hits": 28}
    for i, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 18)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(len(rows)+1, 2)}"

    nota = wb.create_sheet("como_revisar")
    nota["A1"] = "ok"
    nota["B1"] = "1 = encaixa no nicho (hidrômetro / vazão / pressão / calibração do dataset). 0 = falso positivo. vazio = ainda não viu."
    nota["A2"] = "comentario"
    nota["B2"] = "Se 0: por que (ex. decibelímetro, odonto, só 'registro de preços'). Se 1 e o filtro quase perdeu: anota a palavra que faltou."
    nota["A3"] = "corte"
    nota["B3"] = (meta or {}).get("corte", "")
    nota["A4"] = "coletados"
    nota["B4"] = (meta or {}).get("coletados", "")
    nota["A5"] = "nicho"
    nota["B5"] = len(rows)
    nota["A6"] = "datasets"
    nota["B6"] = ", ".join((meta or {}).get("datasets") or [])
    nota["A7"] = "comando"
    nota["B7"] = "python3 pncp_find.py --horas 24 --json --xlsx este_arquivo.xlsx"
    nota.column_dimensions["A"].width = 14
    nota.column_dimensions["B"].width = 110
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Buscador PNCP (espelho Licita Já): recência + nicho/keyword + UF/cidade/valor."
    )
    ap.add_argument(
        "--dataset",
        action="append",
        dest="datasets",
        help="xlsx do gabarito (repetível) ou pasta. Default: data/nicho/dataset_*.xlsx",
    )
    ap.add_argument("--horas", type=int, default=24)
    ap.add_argument("--json", action="store_true", help="stdout JSON (default para agente)")
    ap.add_argument("--xlsx", help="grava planilha além do JSON")
    ap.add_argument(
        "--abertos",
        action="store_true",
        help="só o que ainda dá pra disputar (encerramento futuro; sem data entra — ex. dispensa)",
    )
    ap.add_argument("--min-score", type=int, default=3)
    ap.add_argument("--pausa", type=float, default=PAUSA, help="segundos entre páginas")
    ap.add_argument(
        "--mods",
        default=",".join(str(m) for m in MODALIDADES),
        help="códigos de modalidade separados por vírgula (default 6,8,7,4,5,9)",
    )
    ap.add_argument(
        "--uf",
        help="UFs separadas por vírgula (ex. SP,PR). Uma UF vai na API PNCP; várias filtram depois.",
    )
    ap.add_argument("--valor-min", type=float, default=None, dest="valor_min")
    ap.add_argument("--valor-max", type=float, default=None, dest="valor_max")
    ap.add_argument(
        "--keyword",
        help="termos extra no objeto, separados por vírgula (AND com o nicho; um termo basta)",
    )
    ap.add_argument("--cidade", help="municípios, vírgula (substring, ex. Itu,Campinas)")
    ap.add_argument(
        "--ordem",
        choices=("score", "pub", "encerramento", "valor", "uf"),
        default="score",
        help="ordenação (Licita Já: registro≈pub, abertura≈encerramento)",
    )
    ap.add_argument(
        "--sem-nicho",
        action="store_true",
        help="não aplica léxico do dataset; só keyword/UF/valor/horas (buscador cru)",
    )
    args = ap.parse_args()
    mods = tuple(int(x) for x in args.mods.split(",") if x.strip())

    arquivos = resolver_datasets(args.datasets)
    textos: list[str] = []
    for ds in arquivos:
        if not ds.exists():
            print(f"dataset não achado: {ds}", file=sys.stderr)
            return 2
        chunk = carregar_dataset(ds)
        print(f"[nicho] {ds.name}: {len(chunk)} textos", file=sys.stderr)
        if len(chunk) == 0:
            print(
                f"[nicho] AVISO {ds.name} veio vazio (export Google com fórmula, sem valor). Ignorado.",
                file=sys.stderr,
            )
        textos.extend(chunk)
    # dedup
    seen = set()
    uniq = []
    for t in textos:
        k = fold(t)
        if k not in seen:
            seen.add(k)
            uniq.append(t)
    textos = uniq
    extras = extra_frases(textos)
    print(f"[nicho] união {len(textos)} textos, {len(extras)} bigramas extra", file=sys.stderr)
    if not args.sem_nicho and not textos:
        print("nenhum texto útil nos datasets", file=sys.stderr)
        return 2
    keywords = [fold(k) for k in (args.keyword or "").split(",") if k.strip()]
    cidades = [fold(c) for c in (args.cidade or "").split(",") if c.strip()]

    agora = datetime.now(TZ)
    corte = agora - timedelta(hours=args.horas)
    ufs = [u.strip().upper() for u in (args.uf or "").split(",") if u.strip()]
    uf_api = ufs[0] if len(ufs) == 1 else None
    bruto = coletar(corte, args.pausa, mods, uf=uf_api)

    hits_out = []
    for it in bruto:
        pub = parse_dt(it.get("dataPublicacaoPncp") or it.get("dataInclusao"))
        if not pub or pub < corte:
            continue
        obj = it.get("objetoCompra") or ""
        if args.sem_nicho:
            sc, hits = 0, []
        else:
            sc, hits = pontuar(obj, extras)
            if sc < args.min_score or hits == ["bloqueio"]:
                continue
        if keywords:
            fo = fold(obj)
            if not any(k in fo for k in keywords):
                continue
            hits = list(hits) + [f"kw:{k}" for k in keywords if k in fo]
        if args.abertos:
            enc = parse_dt(it.get("dataEncerramentoProposta"))
            if enc is not None and enc < agora:
                continue
        u = (it.get("unidadeOrgao") or {}).get("ufSigla") or ""
        if ufs and u.upper() not in ufs:
            continue
        mun = fold((it.get("unidadeOrgao") or {}).get("municipioNome") or "")
        if cidades and not any(c in mun for c in cidades):
            continue
        val = it.get("valorTotalEstimado")
        if args.valor_min is not None:
            if not isinstance(val, (int, float)) or val < args.valor_min:
                continue
        if args.valor_max is not None:
            if isinstance(val, (int, float)) and val > args.valor_max:
                continue
        hits_out.append(enriquece(it, sc, hits, pub))
    def _ord(r: dict):
        if args.ordem == "pub":
            return r.get("data_publicacao") or ""
        if args.ordem == "encerramento":
            return r.get("data_encerramento") or "9999"
        if args.ordem == "valor":
            v = r.get("valor_estimado")
            return -(v if isinstance(v, (int, float)) else -1)
        if args.ordem == "uf":
            return r.get("uf") or ""
        return -r["score"]

    reverse = args.ordem in ("score", "valor")
    hits_out.sort(key=_ord, reverse=reverse)

    payload = {
        "corte": corte.isoformat(),
        "coletados": len(bruto),
        "nicho": len(hits_out),
        "datasets": [str(p) for p in arquivos],
        "itens": hits_out,
    }
    if args.xlsx:
        gravar_xlsx(hits_out, Path(args.xlsx).expanduser(), meta=payload)
        print(f"[xlsx] {args.xlsx}", file=sys.stderr)

    # agentes leem stdout; logs vão em stderr
    json.dump(payload, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
